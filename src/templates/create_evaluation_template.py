#!/usr/bin/env python3
"""
Script to create an Excel template for chatbot evaluation criteria.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_questions_sheet(wb):
    """Create the Questions sheet with sample data."""
    ws = wb.create_sheet("Questions", 0)
    
    # Define headers
    headers = [
        "question_id", 
        "category", 
        "question", 
        "expected_answer", 
        "context",
        "difficulty"
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sample_data = [
        {
            "question_id": "Q001",
            "category": "General Knowledge",
            "question": "What is the capital of France?",
            "expected_answer": "The capital of France is Paris.",
            "context": "",
            "difficulty": "Easy"
        },
        {
            "question_id": "Q002",
            "category": "Technical",
            "question": "Explain how a transformer model works in simple terms.",
            "expected_answer": "A transformer model processes text by using attention mechanisms to understand relationships between words. It analyzes all words in a sequence simultaneously rather than sequentially, allowing it to better capture context and meaning.",
            "context": "",
            "difficulty": "Medium"
        },
        {
            "question_id": "Q003",
            "category": "Reasoning",
            "question": "If a train travels at 60 mph, how long will it take to travel 240 miles?",
            "expected_answer": "It will take 4 hours to travel 240 miles at 60 mph.",
            "context": "",
            "difficulty": "Medium"
        }
    ]
    
    # Add sample data
    for row_num, data in enumerate(sample_data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num).value = data.get(header, "")
    
    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        if header in ["question", "expected_answer", "context"]:
            ws.column_dimensions[column_letter].width = 50
        else:
            ws.column_dimensions[column_letter].width = 15
    
    return ws

def create_criteria_sheet(wb):
    """Create the Evaluation Criteria sheet with sample data."""
    ws = wb.create_sheet("Evaluation_Criteria", 1)
    
    # Define headers
    headers = [
        "criteria_id", 
        "name", 
        "description", 
        "weight", 
        "min_score",
        "max_score",
        "passing_threshold"
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sample_data = [
        {
            "criteria_id": "C001",
            "name": "Factual Accuracy",
            "description": "The response contains factually correct information that aligns with the expected answer.",
            "weight": 0.25,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        },
        {
            "criteria_id": "C002",
            "name": "Completeness",
            "description": "The response addresses all aspects of the question without missing key information.",
            "weight": 0.20,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        },
        {
            "criteria_id": "C003",
            "name": "Relevance",
            "description": "The response is directly relevant to the question asked without unnecessary tangents.",
            "weight": 0.15,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        },
        {
            "criteria_id": "C004",
            "name": "Clarity",
            "description": "The response is clear, well-structured, and easy to understand.",
            "weight": 0.15,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        },
        {
            "criteria_id": "C005",
            "name": "Conciseness",
            "description": "The response is appropriately concise without being too brief or too verbose.",
            "weight": 0.10,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        },
        {
            "criteria_id": "C006",
            "name": "Helpfulness",
            "description": "The response is helpful and provides value to the user.",
            "weight": 0.15,
            "min_score": 0,
            "max_score": 5,
            "passing_threshold": 3
        }
    ]
    
    # Add sample data
    for row_num, data in enumerate(sample_data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num).value = data.get(header, "")
    
    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "description":
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = 15
    
    return ws

def create_ragas_metrics_sheet(wb):
    """Create the RAGAS Metrics sheet with sample data."""
    ws = wb.create_sheet("RAGAS_Metrics", 2)
    
    # Define headers
    headers = [
        "metric_id", 
        "name", 
        "description", 
        "weight", 
        "passing_threshold"
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sample_data = [
        {
            "metric_id": "R001",
            "name": "faithfulness",
            "description": "Measures if the generated answer contains information that is contradictory to the context.",
            "weight": 0.25,
            "passing_threshold": 0.7
        },
        {
            "metric_id": "R002",
            "name": "answer_relevancy",
            "description": "Measures if the answer is relevant to the question asked.",
            "weight": 0.25,
            "passing_threshold": 0.7
        },
        {
            "metric_id": "R003",
            "name": "context_relevancy",
            "description": "Measures if the retrieved context is relevant to the question.",
            "weight": 0.20,
            "passing_threshold": 0.7
        },
        {
            "metric_id": "R004",
            "name": "context_precision",
            "description": "Measures the precision of the retrieved context.",
            "weight": 0.15,
            "passing_threshold": 0.7
        },
        {
            "metric_id": "R005",
            "name": "context_recall",
            "description": "Measures the recall of the retrieved context.",
            "weight": 0.15,
            "passing_threshold": 0.7
        }
    ]
    
    # Add sample data
    for row_num, data in enumerate(sample_data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num).value = data.get(header, "")
    
    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "description":
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = 15
    
    return ws

def create_thresholds_sheet(wb):
    """Create the Thresholds sheet with sample data."""
    ws = wb.create_sheet("Thresholds", 3)
    
    # Define headers
    headers = [
        "threshold_id", 
        "name", 
        "description", 
        "value"
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sample_data = [
        {
            "threshold_id": "T001",
            "name": "overall_passing_score",
            "description": "The minimum weighted average score required to pass the evaluation.",
            "value": 0.7
        },
        {
            "threshold_id": "T002",
            "name": "min_criteria_pass_count",
            "description": "The minimum number of criteria that must pass their individual thresholds.",
            "value": 4
        },
        {
            "threshold_id": "T003",
            "name": "min_ragas_score",
            "description": "The minimum RAGAS composite score required to pass.",
            "value": 0.65
        },
        {
            "threshold_id": "T004",
            "name": "factual_accuracy_minimum",
            "description": "The minimum score required specifically for factual accuracy (critical criterion).",
            "value": 3
        }
    ]
    
    # Add sample data
    for row_num, data in enumerate(sample_data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num).value = data.get(header, "")
    
    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "description":
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = 15
    
    return ws

def create_instructions_sheet(wb):
    """Create the Instructions sheet."""
    ws = wb.create_sheet("Instructions", 4)
    
    # Add title
    ws['A1'] = "Chatbot Evaluation Template Instructions"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Add instructions
    instructions = [
        "This Excel template is used to configure the chatbot evaluation framework.",
        "",
        "Sheets:",
        "1. Questions - Define the test questions, expected answers, and related metadata.",
        "2. Evaluation_Criteria - Define the criteria used by the Bedrock LLM to evaluate responses.",
        "3. RAGAS_Metrics - Configure RAGAS evaluation metrics and their weights.",
        "4. Thresholds - Set passing thresholds for the overall evaluation.",
        "",
        "Usage Instructions:",
        "- Modify the sample data with your own questions and criteria.",
        "- Ensure all question_ids and criteria_ids are unique.",
        "- The sum of weights in the Evaluation_Criteria sheet should equal 1.0.",
        "- The sum of weights in the RAGAS_Metrics sheet should equal 1.0.",
        "- Do not change the column headers or sheet names as they are used by the evaluation framework.",
        "",
        "For more information, refer to the documentation in the /docs directory."
    ]
    
    for i, line in enumerate(instructions, 3):
        ws[f'A{i}'] = line
        if line and not line.startswith(" "):
            ws[f'A{i}'].font = Font(bold=True)
        if line.startswith("  "):
            ws.merge_cells(f'A{i}:F{i}')
    
    # Adjust column width
    ws.column_dimensions['A'].width = 100
    
    return ws

def create_evaluation_template():
    """Create the evaluation template Excel file."""
    output_dir = os.path.join(os.getcwd(), "data")
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "chatbot_evaluation_template.xlsx")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    create_questions_sheet(wb)
    create_criteria_sheet(wb)
    create_ragas_metrics_sheet(wb)
    create_thresholds_sheet(wb)
    create_instructions_sheet(wb)
    
    # Save workbook
    wb.save(output_file)
    print(f"Evaluation template created at: {output_file}")
    return output_file

if __name__ == "__main__":
    create_evaluation_template()
